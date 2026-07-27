#!/usr/bin/env python3
"""Read an Excel workbook and output sheets as Markdown tables.

Usage:
    excel_read.py <file.xlsx> [--sheet NAME] [--columns A,B,J] [--filter COL=VALUE]

Options:
    --sheet NAME      Read only this sheet (default: all sheets)
    --columns A,B,J   Only include these columns (letter or 1-based index)
    --filter COL=VAL  Only show rows where column COL contains VAL (case-insensitive)
    --csv             Output as CSV instead of Markdown table
"""

import sys
import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def col_letter_to_index(col_spec):
    """Convert column letter (A, B, J) or 1-based number to 0-based index."""
    try:
        return int(col_spec) - 1
    except ValueError:
        return column_index_from_string(col_spec.upper()) - 1


def read_sheet(ws, col_indices=None, filter_col=None, filter_val=None):
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c) if c is not None else "" for c in row]
        rows.append(cells)

    if not rows:
        return []

    # Determine max columns
    max_cols = max(len(r) for r in rows)
    # Pad rows
    rows = [r + [""] * (max_cols - len(r)) for r in rows]

    # Filter columns
    if col_indices:
        valid = [i for i in col_indices if i < max_cols]
        rows = [[r[i] for i in valid] for r in rows]

    # Filter rows
    if filter_col is not None and filter_val is not None:
        header = rows[0]
        # Find filter column index
        fi = None
        try:
            fi = col_letter_to_index(filter_col)
            # Adjust if we've already filtered columns
            if col_indices:
                fi = col_indices.index(fi) if fi in col_indices else None
        except (ValueError, IndexError):
            # Try by header name
            for i, h in enumerate(header):
                if h.lower() == filter_col.lower():
                    fi = i
                    break

        if fi is not None and fi < len(header):
            filtered = [rows[0]]  # keep header
            for r in rows[1:]:
                if fi < len(r) and filter_val.lower() in r[fi].lower():
                    filtered.append(r)
            rows = filtered

    return rows


def rows_to_markdown(rows, sheet_name=None):
    if not rows:
        return ""

    lines = []
    if sheet_name:
        lines.append(f"## {sheet_name}")
        lines.append("")

    # Header
    lines.append("| " + " | ".join(rows[0]) + " |")
    lines.append("| " + " | ".join("---" for _ in rows[0]) + " |")
    for row in rows[1:]:
        lines.append("| " + " | ".join(row) + " |")

    return "\n".join(lines)


def rows_to_csv(rows):
    import csv
    import io
    output = io.StringIO()
    writer = csv.writer(output)
    for row in rows:
        writer.writerow(row)
    return output.getvalue()


def main():
    parser = argparse.ArgumentParser(description="Read Excel files")
    parser.add_argument("file", help="Path to .xlsx file")
    parser.add_argument("--sheet", help="Sheet name to read (default: all)")
    parser.add_argument("--columns", help="Comma-separated column letters or numbers (e.g. A,B,J)")
    parser.add_argument("--filter", dest="filter_expr", help="Filter: COL=VALUE")
    parser.add_argument("--csv", action="store_true", help="Output as CSV")
    args = parser.parse_args()

    wb = load_workbook(args.file, read_only=True, data_only=True)

    col_indices = None
    if args.columns:
        col_indices = [col_letter_to_index(c.strip()) for c in args.columns.split(",")]

    filter_col = filter_val = None
    if args.filter_expr and "=" in args.filter_expr:
        filter_col, filter_val = args.filter_expr.split("=", 1)

    sheets = [args.sheet] if args.sheet else wb.sheetnames

    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}", file=sys.stderr)
            continue

        ws = wb[sheet_name]
        rows = read_sheet(ws, col_indices, filter_col, filter_val)

        if args.csv:
            print(rows_to_csv(rows))
        else:
            print(rows_to_markdown(rows, sheet_name if len(sheets) > 1 else None))
            print()

    wb.close()


if __name__ == "__main__":
    main()
